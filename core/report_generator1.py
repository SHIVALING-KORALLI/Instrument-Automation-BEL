# report_generator.py
import os
import datetime
from typing import List, Dict
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ReportGenerator:
    """Generates structured Excel reports with DTRC board/channel organization."""
    
    def __init__(self, report_dir="reports"):
        self.report_dir = report_dir
        os.makedirs(report_dir, exist_ok=True)
        
        # Create unique filename per run with timestamp
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        self.filepath = os.path.join(report_dir, f"report_{timestamp}.xlsx")
        
        # Create new workbook
        self.wb = Workbook()
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])
        
        # Track channel positions per board: {board_no: {channel_no: col_start}}
        self.board_channel_positions = {}

    def add_dtrc_results(self, board_no: int, channel_no: int, results: List[Dict]):
        """
        Add results for a specific DTRC board and channel.
        
        Layout:
        - One sheet per board: "DTRC Board {board_no}"
        - Channels arranged horizontally (side-by-side)
        - Each channel has: header, data table, chart, insights
        """
        if not results:
            return
        
        sheet_name = f"DTRC Board {board_no}"
        
        # Get or create sheet for this board
        if sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
        else:
            ws = self.wb.create_sheet(sheet_name)
            # Add yellow board header at the top
            self._add_board_header(ws, board_no)
            # Initialize position tracking for this board
            self.board_channel_positions[board_no] = {}
        
        # Determine column position for this channel
        col_start = self._get_channel_column(board_no, channel_no)
        
        # Add channel section
        self._add_channel_section(ws, channel_no, results, col_start)
        
        # Store position
        self.board_channel_positions[board_no][channel_no] = col_start

    def _add_board_header(self, ws, board_no: int):
        """Add yellow merged header for the board at top of sheet."""
        ws.merge_cells("A1:F1")
        cell = ws["A1"]
        cell.value = f"DTRC Board {board_no}"
        cell.font = Font(bold=True, size=14, color="000000")
        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

    def _get_channel_column(self, board_no: int, channel_no: int) -> int:
        """Determine starting column for a channel (1-indexed)."""
        if board_no not in self.board_channel_positions:
            self.board_channel_positions[board_no] = {}
        
        existing_channels = self.board_channel_positions[board_no]
        
        if channel_no in existing_channels:
            # Overwrite existing channel
            return existing_channels[channel_no]
        
        if not existing_channels:
            # First channel starts at column A (1)
            return 1
        
        # Place after the last channel with 3-column gap
        max_col = max(existing_channels.values())
        return max_col + 6  # 3 columns for data + 3 columns gap

    def _add_channel_section(self, ws, channel_no: int, results: List[Dict], col_start: int):
        """
        Add complete channel section: header, data, chart, insights.
        
        Args:
            ws: worksheet
            channel_no: channel number
            results: list of measurement results
            col_start: starting column (1-indexed)
        """
        start_row = 3  # Row 1 is board header, row 2 is blank
        
        # === CHANNEL HEADER ===
        header_row = start_row
        ws.merge_cells(
            start_row=header_row, 
            start_column=col_start, 
            end_row=header_row, 
            end_column=col_start + 2
        )
        header_cell = ws.cell(row=header_row, column=col_start)
        header_cell.value = f"Channel {channel_no}"
        header_cell.font = Font(bold=True, size=12, color="FFFFFF")
        header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[header_row].height = 22
        
        # === TABLE HEADERS ===
        table_header_row = start_row + 1
        headers = ["Spot", "Freq (Hz)", "Power (dBm)"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=table_header_row, column=col_start + i)
            cell.value = header
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # === DATA ROWS ===
        data_start_row = table_header_row + 1
        for idx, r in enumerate(results):
            row = data_start_row + idx
            
            # Spot (hex)
            cell = ws.cell(row=row, column=col_start)
            cell.value = r["spot"] 
            # cell.value = int(r["spot"], 16)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Frequency
            cell = ws.cell(row=row, column=col_start + 1)
            cell.value = r["freq_hz"]
            cell.number_format = '0.00E+00'  # Scientific notation
            cell.alignment = Alignment(horizontal="right")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Power
            cell = ws.cell(row=row, column=col_start + 2)
            cell.value = r["power_dbm"]
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal="right")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        data_end_row = data_start_row + len(results) - 1
        
        # Auto-adjust column widths
        ws.column_dimensions[get_column_letter(col_start)].width = 10
        ws.column_dimensions[get_column_letter(col_start + 1)].width = 15
        ws.column_dimensions[get_column_letter(col_start + 2)].width = 14
        
        # === CHART ===
        chart_start_row = data_end_row + 3
        self._add_chart(ws, channel_no, col_start, data_start_row, data_end_row, chart_start_row)
        
        # === INSIGHTS ===
        insights_start_row = chart_start_row + 18
        self._add_insights(ws, results, col_start, insights_start_row)

    def _add_chart(self, ws, channel_no: int, col_start: int, 
                   data_start_row: int, data_end_row: int, chart_row: int):
        """Add scatter chart for Power vs Spot."""
        chart = ScatterChart()
        chart.title = f"Power vs Spot - Channel {channel_no}"
        chart.style = 13
        chart.x_axis.title = "Spot (Decimal)"
        chart.y_axis.title = "Power (dBm)"
        chart.height = 10  # Chart height in cm
        chart.width = 15   # Chart width in cm
        
        # Add gridlines for better readability
        chart.x_axis.majorGridlines = None
        chart.y_axis.majorGridlines = None
        
        # Convert hex spots to decimal for chart x-axis
        # We need to create a helper column with decimal values
        decimal_col = col_start + 3  # One column to the right of Power
        
        # Add decimal spot values in hidden column
        for idx in range(data_start_row, data_end_row + 1):
            spot_hex = ws.cell(row=idx, column=col_start).value
            decimal_val = int(spot_hex, 16)
            ws.cell(row=idx, column=decimal_col).value = decimal_val
        
        # Hide the decimal column
        ws.column_dimensions[get_column_letter(decimal_col)].hidden = True
        
        # Define data references
        # X values: decimal spots (hidden column)
        xvalues = Reference(ws, min_col=decimal_col, min_row=data_start_row, max_row=data_end_row)
        # Y values: power (dBm)
        yvalues = Reference(ws, min_col=col_start + 2, min_row=data_start_row, max_row=data_end_row)
        
        # Create series
        from openpyxl.chart import Series
        series = Series(yvalues, xvalues, title=f"Channel {channel_no}")
        chart.series.append(series)
        
        # Position chart
        chart_cell = ws.cell(row=chart_row, column=col_start)
        ws.add_chart(chart, chart_cell.coordinate)

    def _add_insights(self, ws, results: List[Dict], col_start: int, start_row: int):
        """Add insights section below the chart."""
        if not results:
            return
        
        # Calculate statistics
        powers = [r["power_dbm"] for r in results]
        spots_hex = [r["spot"] for r in results]
        spots_dec = [int(s, 16) for s in spots_hex]
        
        max_power = round(max(powers), 2)
        min_power = round(min(powers), 2)
        avg_power = round(sum(powers) / len(powers), 2)
        max_spot_idx = powers.index(max(powers))
        min_spot_idx = powers.index(min(powers))
        max_spot = spots_dec[max_spot_idx]
        min_spot = spots_dec[min_spot_idx]
        power_range = round(max_power - min_power, 2)
        
        # Insights title
        title_cell = ws.cell(row=start_row, column=col_start)
        title_cell.value = "Insights"
        title_cell.font = Font(bold=True, size=11, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Merge title across 2 columns
        ws.merge_cells(
            start_row=start_row,
            start_column=col_start,
            end_row=start_row,
            end_column=col_start + 1
        )
        
        # Insights data
        insights = [
            ("Max Power (dBm)", max_power),
            ("Min Power (dBm)", min_power),
            ("Average Power (dBm)", avg_power),
            ("Spot @ Max Power", f"{max_spot} (0x{spots_hex[max_spot_idx]})"),
            ("Spot @ Min Power", f"{min_spot} (0x{spots_hex[min_spot_idx]})"),
            ("Power Range (dB)", power_range),
        ]
        
        current_row = start_row + 1
        for label, value in insights:
            # Label cell
            label_cell = ws.cell(row=current_row, column=col_start)
            label_cell.value = label
            label_cell.font = Font(bold=True, size=9)
            label_cell.alignment = Alignment(horizontal="left")
            label_cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Value cell
            value_cell = ws.cell(row=current_row, column=col_start + 1)
            value_cell.value = value
            value_cell.alignment = Alignment(horizontal="center")
            value_cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            current_row += 1

    def save(self):
        """Save the workbook and return file path."""
        self.wb.save(self.filepath)
        print(f"✅ Report saved: {self.filepath}")
        return self.filepath
