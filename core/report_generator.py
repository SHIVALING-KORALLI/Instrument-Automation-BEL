import os
import datetime
from typing import List, Dict
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ReportGenerator:
    """Generates structured Excel reports with DTRC board/channel organization."""

    def __init__(self, report_dir="reports"):
        self.report_dir = report_dir
        os.makedirs(report_dir, exist_ok=True)

        today = datetime.date.today().strftime("%Y%m%d")
        self.filepath = os.path.join(report_dir, f"report_{today}.xlsx")

        try:
            if os.path.exists(self.filepath):
                self.wb = load_workbook(self.filepath)
                print(f"📂 Loaded existing daily report: {self.filepath}")
            else:
                self.wb = Workbook()
                if "Sheet" in self.wb.sheetnames:
                    self.wb.remove(self.wb["Sheet"])
                print(f"🆕 Created new daily report: {self.filepath}")
        except Exception as e:
            raise RuntimeError(f"❌ Error initializing workbook: {e}")

        self.board_channel_positions = self._scan_existing_positions()

    def _scan_existing_positions(self):
        """Rebuild channel position map from existing workbook."""
        mapping = {}
        try:
            for ws in self.wb.worksheets:
                if not ws.title.startswith("DTRC Board "):
                    continue
                try:
                    board_no = int(ws.title.replace("DTRC Board ", "").strip())
                except ValueError:
                    continue
                mapping[board_no] = {}
                for row in ws.iter_rows(min_row=3, max_row=4, max_col=50):
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith("Channel "):
                            try:
                                ch_no = int(cell.value.replace("Channel ", "").strip())
                                mapping[board_no][ch_no] = cell.column
                            except ValueError:
                                pass
        except Exception as e:
            print(f"⚠️ Error scanning existing positions: {e}")
        return mapping

    def add_dtrc_results(self, board_no: int, channel_no: int, results: List[Dict]):
        """Add results for a specific DTRC board and channel."""
        if not results:
            raise ValueError("No results provided for report generation")

        try:
            sheet_name = f"DTRC Board {board_no}"
            if sheet_name in self.wb.sheetnames:
                ws = self.wb[sheet_name]
            else:
                ws = self.wb.create_sheet(sheet_name)
                self._add_board_header(ws, board_no)
                self.board_channel_positions[board_no] = {}

            col_start = self._get_channel_column(board_no, channel_no)
            self._add_channel_section(ws, channel_no, results, col_start)
            self.board_channel_positions[board_no][channel_no] = col_start
        except Exception as e:
            raise RuntimeError(f"❌ Error adding DTRC results (Board {board_no}, Channel {channel_no}): {e}")

    def _add_board_header(self, ws, board_no: int):
        try:
            ws.merge_cells("A1:F1")
            cell = ws["A1"]
            cell.value = f"DTRC Board {board_no}"
            cell.font = Font(bold=True, size=14, color="000000")
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 25
        except Exception as e:
            raise RuntimeError(f"❌ Error adding board header for Board {board_no}: {e}")

    def _get_channel_column(self, board_no: int, channel_no: int) -> int:
        """Determine starting column for a channel (1-indexed)."""
        if board_no not in self.board_channel_positions:
            self.board_channel_positions[board_no] = {}

        existing = self.board_channel_positions[board_no]
        if channel_no in existing:
            return existing[channel_no]
        return 1 if not existing else max(existing.values()) + 6

    def _add_channel_section(self, ws, channel_no: int, results: List[Dict], col_start: int):
        """Add complete channel section: data, chart, insights."""
        try:
            start_row = 3
            header_row = start_row

            # --- Channel header ---
            ws.merge_cells(start_row=header_row, start_column=col_start, end_row=header_row, end_column=col_start + 2)
            header_cell = ws.cell(row=header_row, column=col_start)
            header_cell.value = f"Channel {channel_no}"
            header_cell.font = Font(bold=True, size=12, color="FFFFFF")
            header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_cell.alignment = Alignment(horizontal="center", vertical="center")

            # --- Table headers ---
            table_header_row = start_row + 1
            headers = ["Spot (Deci)", "Freq (Hz)", "Power (dBm)"]
            for i, header in enumerate(headers):
                cell = ws.cell(row=table_header_row, column=col_start + i)
                cell.value = header
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # --- Data rows ---
            data_start_row = table_header_row + 1
            for idx, r in enumerate(results):
                row = data_start_row + idx
                try:
                    spot_val = int(r["spot"], 16)
                except Exception:
                    spot_val = r["spot"]  # fallback if already int

                ws.cell(row=row, column=col_start, value=spot_val)
                ws.cell(row=row, column=col_start + 1, value=r.get("freq_hz", 0)).number_format = "0"
                ws.cell(row=row, column=col_start + 2, value=r.get("power_dbm", 0.0)).number_format = "0.00"

            data_end_row = data_start_row + len(results) - 1
            ws.column_dimensions[get_column_letter(col_start)].width = 12
            ws.column_dimensions[get_column_letter(col_start + 1)].width = 15
            ws.column_dimensions[get_column_letter(col_start + 2)].width = 14

            # --- Chart ---
            chart_start_row = data_end_row + 3
            self._add_chart(ws, channel_no, col_start, data_start_row, data_end_row, chart_start_row)

            # --- Insights ---
            self._add_insights(ws, results, col_start, chart_start_row + 20)

        except Exception as e:
            raise RuntimeError(f"❌ Error creating section for Channel {channel_no}: {e}")

    def _add_chart(self, ws, channel_no: int, col_start: int, data_start_row: int, data_end_row: int, chart_row: int):
        """Add scatter chart for Power vs Spot with labeled data points."""
        try:
            chart = ScatterChart()
            chart.title = f"Power vs Spot - Channel {channel_no}"
            chart.style = 13
            chart.x_axis.title = "Spot (Decimal)"
            chart.y_axis.title = "Power (dBm)"
            chart.height = 10
            chart.width = 15

            xvalues = Reference(ws, min_col=col_start, min_row=data_start_row, max_row=data_end_row)
            yvalues = Reference(ws, min_col=col_start + 2, min_row=data_start_row, max_row=data_end_row)

            series = Series(yvalues, xvalues, title=f"Channel {channel_no}")
            series.dLbls = DataLabelList()
            series.dLbls.showVal = True
            chart.series.append(series)
            ws.add_chart(chart, f"{get_column_letter(col_start)}{chart_row}")

        except Exception as e:
            raise RuntimeError(f"❌ Error adding chart for Channel {channel_no}: {e}")

    def _add_insights(self, ws, results: List[Dict], col_start: int, start_row: int):
        """Add computed summary stats below chart."""
        try:
            powers = [r["power_dbm"] for r in results]
            spots_dec = [int(r["spot"], 16) if isinstance(r["spot"], str) else r["spot"] for r in results]

            max_power = max(powers)
            min_power = min(powers)
            avg_power = round(sum(powers) / len(powers), 2)
            max_spot = spots_dec[powers.index(max_power)]
            min_spot = spots_dec[powers.index(min_power)]
            power_range = round(max_power - min_power, 2)

            ws.merge_cells(start_row=start_row, start_column=col_start, end_row=start_row, end_column=col_start + 1)
            title = ws.cell(row=start_row, column=col_start)
            title.value = "Insights"
            title.font = Font(bold=True, size=11, color="FFFFFF")
            title.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            title.alignment = Alignment(horizontal="center")

            insights = [
                ("Max Power (dBm)", max_power),
                ("Min Power (dBm)", min_power),
                ("Average Power (dBm)", avg_power),
                ("Spot @ Max Power", max_spot),
                ("Spot @ Min Power", min_spot),
                ("Power Range (dB)", power_range),
            ]
            for i, (label, value) in enumerate(insights, start=start_row + 1):
                ws.cell(row=i, column=col_start, value=label).font = Font(bold=True, size=9)
                ws.cell(row=i, column=col_start + 1, value=value)
        except Exception as e:
            raise RuntimeError(f"❌ Error adding insights section: {e}")

    def save(self):
        try:
            self.wb.save(self.filepath)
            print(f"✅ Report saved successfully: {self.filepath}")
            return self.filepath
        except Exception as e:
            raise RuntimeError(f"❌ Error saving report: {e}")
